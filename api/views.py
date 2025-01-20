from .models import Client,User,Pharmacie,Commande,Conseil, Recherche, Notification, Invoice, InvoiceItem, InvoicePayment, WalletPharmacie,  WalletPharmacieHistory
from .serializers import ClientSerializer, UserLoginSerializer,PharmacieSerializer, UserSerializer, CommandeSerializer,ConseilSerializer, RechercheSerializer,  NotificationSerializer, InvoiceSerializer, InvoicePaymentSerializer, WalletPharmacieSerializer,  WalletPharmacieHistorySerializer,  TransactionSerializer
from rest_framework.views import APIView
from rest_framework.generics import RetrieveAPIView
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import AuthenticationFailed
from .permissions import IsPharmacieOrReadOnly,IsSuperUserOrReadOnly,IsClientOrReadOnly,IsPharmacieCanModifyCommande, IsPharmacieOrClient
from .backends import EmailBackend
from .utils import has_key_client,has_key_pharmacie, generer_code, send_notification, generate_reference, send_sms, send_sms_jetfy,  upload_to_s3
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.http import Http404
from rest_framework import serializers
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q
from django.core.serializers import serialize
from rest_framework import generics
import jwt
import re
import os
import datetime
from rest_framework_simplejwt.exceptions import TokenError
from rest_framework import status
from django.db import transaction
from decimal import Decimal
from .utils import upload_to_s3
from rest_framework.parsers import MultiPartParser, FormParser
import openpyxl




# from .utils import pharmacies_within_radius


class ClientRegistrationAPIView(APIView):
    serializer_class = ClientSerializer

    example_request = {
        "user": {
            "email": "client@example.com",
            "username": "votre nom de famille",
            "password": "motdepasseclient",
            "is_pharmacy": False
        },
        "prenom": "John",
        "nom": "Light",
        "adresse": "123 Rue de la Ville",
        "ville": "Ville",
        "phone": "0123456789",
        "image": "lien_de_l_image.jpg",
        "n_cmu": "CMU123",
        "n_assurance": "Assurance123",
        "sexe": "Homme",
        "maladie_chronique": "Aucune",
        "poids": 75.5,
        "taille": 180.0
    }

    @swagger_auto_schema(
        request_body=ClientSerializer,
        manual_parameters=[
            openapi.Parameter(
                'example',
                openapi.IN_QUERY,
                description='Exemple de JSON pour l\'enregistrement',
                type=openapi.TYPE_STRING,
                example=example_request,
            ),
        ],
        responses={
            201: openapi.Response('Client enregistré avec succès', examples={
                'application/json': {
                    'detail': 'Client enregistré avec succès!',
                },
            }),
            400: openapi.Response('Erreur de validation', examples={
                'application/json': {
                    'error': 'Détails sur l\'erreur de validation',
                },
            }),
        },
        operation_summary='Enregistrement du client',
        operation_description='Cette vue vous permet d\'enregistrer un nouveau client.',
    )
    
    def post(self, request):
        try:
            serializer = self.serializer_class(data=request.data)
            if serializer.is_valid(raise_exception=False):
                # Créez et enregistrez un nouvel utilisateur
                numero=serializer.validated_data['num_pharmacie']

                if numero:
                    pharmacie=Pharmacie.objects.filter(num_pharmacie=numero).first()
                    
                    if  not pharmacie:
                        return Response({"detail": "Désolé le numero de la pharmacie est incorrecte."}, status=status.HTTP_400_BAD_REQUEST)
                    
                new_user = User.objects.create_user(
                    email=serializer.validated_data['user']['email'],
                    username=serializer.validated_data['user']['username'],
                    password=request.data['user']['password'],
                    is_pharmacie=False
                )

                # Associez le nouvel utilisateur au modèle Client
                    
                new_client = Client.objects.create(
                    user=new_user,
                    prenom=serializer.validated_data['prenom'],
                    nom=serializer.validated_data.get('nom', ''),
                    adresse=serializer.validated_data['adresse'],
                    ville=serializer.validated_data['ville'],
                    phone=serializer.validated_data['phone'],
                    image=serializer.validated_data['image'],
                    n_cmu=serializer.validated_data['n_cmu'],
                    n_assurance=serializer.validated_data['n_assurance'],
                    sexe=serializer.validated_data['sexe'],
                    maladie_chronique=serializer.validated_data['maladie_chronique'],
                    poids=serializer.validated_data['poids'],
                    taille=serializer.validated_data['taille'],
                    num_pharmacie=serializer.validated_data['num_pharmacie']
                )

                #Send SMS
                send_notification = send_sms_jetfy(new_client.phone,f"Bonjour {new_client.prenom}, votre compte a bien été créé. Vous pouvez vous connecter sur l'application mobile.")
                print("send_notification: ",send_notification)
                # Utilisez les tokens pour générer les cookies
                refresh = RefreshToken.for_user(new_user)
                data = {
                        'refresh': str(refresh),
                        'access': str(refresh.access_token),
                        'detail': f"Client {new_client.prenom} enregistré avec succès!",
                        'user_data':ClientSerializer(new_client,many=False).data
                    }
                response = Response(data, status=status.HTTP_201_CREATED)

                return response
                
            list_erreur=has_key_client(serializer.errors)
            return Response(list_erreur, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ClientUpdateAPIView(APIView):
    authentication_classes = (JWTAuthentication,)
    
    permission_classes = (IsClientOrReadOnly,)
    
    def put(self, request):
        # Créez et enregistrez un nouvel utilisateur
        # numero= request.data['num_pharmacie']
        try:

            formData = request.data
            
            if 'num_pharmacie'  in request.data and request.data['num_pharmacie']:
                pharmacie=Pharmacie.objects.filter(num_pharmacie=request.data['num_pharmacie']).first()
            
                if  not pharmacie:
                    return Response({"detail": "Désolé, aucune pharmacie trouvée avec ce identifiant."}, status=status.HTTP_400_BAD_REQUEST)
                
            user = request.user
            if not user:
                return Response({'detail': "Désolé, vous devez être authentifié pour effectuer cette action."}, status=status.HTTP_400_BAD_REQUEST)
            
            # Modifier mot de passe User
            if 'user' in request.data and 'password' in request.data['user']:
                user.set_password(request.data['user']['password'])
                user.save()
                formData.pop('user')
                
            
            client = request.user.client_user
            
            for key in formData:
                setattr(client, key, formData[key])
                    
            client.save()
                

            # Utilisez les tokens pour générer les cookies
            refresh = RefreshToken.for_user(user)
            data = {
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                    'detail': f"Client {client.prenom} modifié avec succès!",
                    'user_data':ClientSerializer(client,many=False).data
                }
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ClientDetailAPIView(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = [IsAuthenticated]
    
    def get(self,request, pk):
        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            raise Http404
        
        return Response(ClientSerializer(client,many=False).data, status=status.HTTP_200_OK)

class PharmacieRegistrationAPIView(APIView):

    serializer_class = PharmacieSerializer
    example_request = {
        "user": {
            "email": "pharmacie@example.com",
            "username": "nom du docteur de la pharmacie",
            "password": "motdepassepharmacie",
            "is_pharmacie": True
        },
        "num_pharmacie": "12345",
        "nom_pharmacie": "Pharmacie ABC",
        "adresse_pharmacie": "456 Rue de la Pharmacie",
        "commune_pharmacie": "Commune",
        "ville_pharmacie": "Ville",
        "numero_contact_pharmacie": "0123456789",
        "horaire_ouverture_pharmacie": "9h00 - 18h00"
    }

    @swagger_auto_schema(
        request_body=PharmacieSerializer,
        manual_parameters=[
            openapi.Parameter(
                'example',
                openapi.IN_QUERY,
                description='Exemple de JSON pour l\'enregistrement de la pharmacie',
                type=openapi.TYPE_STRING,
                example=example_request,
            ),
        ],
        responses={
            201: openapi.Response('Pharmacie enregistrée avec succès', examples={
                'application/json': {
                    'detail': 'Pharmacie enregistrée avec succès!',
                },
            }),
            400: openapi.Response('Erreur de validation', examples={
                'application/json': {
                    'error': 'Détails sur l\'erreur de validation',
                },
            }),
        },
        operation_summary='Enregistrement de la pharmacie',
        operation_description='Cette vue vous permet d\'enregistrer une nouvelle pharmacie.',
    )
    
    def post(self, request):
        try:

            serializer = self.serializer_class(data=request.data)
            if serializer.is_valid(raise_exception=False):
                # Créez et enregistrez un nouvel utilisateur
                new_user = User.objects.create_user(
                    email=serializer.validated_data['user']['email'],
                    username=serializer.validated_data['user']['username'],
                    password=request.data['user']['password'],
                    is_pharmacie=True
                )

                # Associez le nouvel utilisateur au modèle Pharmacie
                
                new_pharmacie = Pharmacie.objects.create(
                    user=new_user,
                    num_pharmacie="",
                    nom_pharmacie=serializer.validated_data['nom_pharmacie'],
                    adresse_pharmacie=serializer.validated_data['adresse_pharmacie'],
                    commune_pharmacie=serializer.validated_data['commune_pharmacie'],
                    ville_pharmacie=serializer.validated_data['ville_pharmacie'],
                    numero_contact_pharmacie=serializer.validated_data['numero_contact_pharmacie'],
                    horaire_ouverture_pharmacie=serializer.validated_data['horaire_ouverture_pharmacie'],
                )
                
                # Enregistrer code pharmacie
                code = generer_code(new_pharmacie.nom_pharmacie, new_pharmacie.pk, longueur=6)
                new_pharmacie.num_pharmacie = code
                new_pharmacie.save()
                
                
                refresh = RefreshToken.for_user(new_user)

                print(f"Pharmacie {new_pharmacie.nom_pharmacie} enregistrée avec succès!")
                
                data = {
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                    'detail': f"Pharmacie {new_pharmacie.nom_pharmacie} enregistrée avec succès!",
                    'user_data':PharmacieSerializer(new_pharmacie,many=False).data
                    
                }
                response = Response(data, status=status.HTTP_201_CREATED)

                return response
            list_erreur=has_key_pharmacie(serializer.errors)
            
            return Response(list_erreur, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class PharmacieUpdateAPIView(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = [IsAuthenticated]

    serializer_class = PharmacieSerializer
    def put(self, request):
        try:

            serializer = self.serializer_class(data=request.data)
            if request.user.is_pharmacie != True:
                return Response({ 'detail': "Vous ne pouvez pas effectuer cette action"}, status=status.HTTP_400_BAD_REQUEST)
            
            # if serializer.is_valid(raise_exception=False):
            #   Modifier et enregistrez un nouvel utilisateur
            user = request.user
            if 'user' in request.data and 'password' in request.data['user']:
                user.set_password(request.data['user']['password'])
                user.save()
            
            # validateData = serializer.validated_data
            # Modifier les information du modèle Pharmacie
            # if 'user' in validateData:
            #      user_data = validateData.pop('user')
                
            pharmacie = request.user.pharmacie_user
            # print(pharmacie)
            # pharmacie.update(**validateData)
            
            pharmacie.nom_pharmacie=request.data['nom_pharmacie']
            pharmacie.adresse_pharmacie=request.data['adresse_pharmacie']
            pharmacie.commune_pharmacie=request.data['commune_pharmacie']
            pharmacie.ville_pharmacie=request.data['ville_pharmacie']
            pharmacie.numero_contact_pharmacie=request.data['numero_contact_pharmacie']
            pharmacie.horaire_ouverture_pharmacie=request.data['horaire_ouverture_pharmacie']
            pharmacie.degarde=request.data['degarde']
            pharmacie.save()
            
                
            refresh = RefreshToken.for_user(user)

            
            data = {
                'refresh': str(refresh),
                'access': str(refresh.access_token),
                'detail': f"Information Pharmacie {pharmacie.nom_pharmacie} modifiée avec succès!",
                'user_data':PharmacieSerializer(pharmacie,many=False).data
                
            }
            response = Response(data, status=status.HTTP_201_CREATED)

            return response
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class PharmacieDetailAPIView(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            pharmacie = Pharmacie.objects.get(pk=pk)
        except Pharmacie.DoesNotExist:
            raise Http404
        
        return Response(PharmacieSerializer(pharmacie,many=False).data, status=status.HTTP_200_OK)

class UserDetailView(RetrieveAPIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = [IsAuthenticated]

    serializer_class = UserSerializer

    @swagger_auto_schema(
        operation_summary='Récupérer les informations de l\'utilisateur connecté',
        operation_description='Récupère les informations de l\'utilisateur connecté, y compris son ID, nom d\'utilisateur et email.',
        responses={
            200: openapi.Response('Informations utilisateur récupérées avec succès', UserSerializer),
            401: 'Non authentifié - L\'utilisateur doit être connecté pour accéder à cette ressource.',
        }
    )
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)


class UserLoginAPIView(APIView):
    authentication_classes = []
    serializer_class = UserLoginSerializer
    permission_classes = (AllowAny,)

    example_request = {
        "email": "pharmacie@example.com",
        "password": "motdepasse",
    }

    @swagger_auto_schema(
        request_body=UserLoginSerializer,
        manual_parameters=[
            openapi.Parameter(
                'example',
                openapi.IN_QUERY,
                description='Exemple de JSON pour la connexion de l\'utilisateur',
                type=openapi.TYPE_STRING,
                example=example_request,
            ),
        ],
        responses={
            200: openapi.Response('Utilisateur connecté avec succès', examples={
                'application/json': {
                    'refresh': 'refresh_token',
                    'access': 'access_token',
                    'detail': 'Utilisateur connecté avec succès',
                },
            }),
            401: openapi.Response('Échec de l\'authentification', examples={
                'application/json': {
                    'error': 'Email ou mot de passe incorrect',
                },
            }),
        },
        operation_summary='Connexion de l\'utilisateur',
        operation_description='Cette vue vous permet de connecter un utilisateur existant.',
    )
    def post(self, request):
        email = request.data.get('email', None)
        user_password = request.data.get('password', None)

        if not user_password or not email:
            raise AuthenticationFailed('L\'email et le mot de passe de l\'utilisateur sont nécessaires.')

        user_instance = EmailBackend().authenticate(request, email=email, password=user_password)

        if user_instance is not None:
            if user_instance.is_active:
                refresh = RefreshToken.for_user(user_instance)
                
                if user_instance.is_pharmacie == True:
                    data = PharmacieSerializer(user_instance.pharmacie_user,many=False).data
                else:
                    data = ClientSerializer(user_instance.client_user,many=False).data
                   
                response_data = {
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                    'detail': 'Utilisateur connecté avec succès',
                    'user_data': data
                }

                return Response(response_data, status=status.HTTP_200_OK)
            else:
                raise AuthenticationFailed('Le compte utilisateur n\'est pas actif.')
        else:
            raise AuthenticationFailed('Email ou mot de passe incorrect.')

class UserByTokenViewAPI(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)

    def post(self, request):

        if request.user.is_pharmacie == True:
            user = PharmacieSerializer(request.user.pharmacie_user,many=False).data
        else:
            user = ClientSerializer(request.user.client_user, many=False).data
            
        # Si le token n'est pas fourni, renvoie une réponse indiquant que l'utilisateur est déjà déconnecté
        return Response({
            "detail": "utilisateur recuperée avec succès",
            'user_data': user
        }, status=status.HTTP_200_OK)   
        
class UserLogoutViewAPI(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)

    @swagger_auto_schema(
        responses={
            200: openapi.Response('Déconnexion réussie', examples={
                'application/json': {
                    'message': 'Déconnexion réussie.',
                },
            }),
            400: openapi.Response('Erreur lors de la déconnexion', examples={
                'application/json': {
                    'error': 'Une erreur est survenue lors de la déconnexion.',
                },
            }),
            401: openapi.Response('Utilisateur déjà déconnecté', examples={
                'application/json': {
                    'message': 'L\'utilisateur est déjà déconnecté.',
                },
            }),
        },
        operation_summary='Déconnexion de l\'utilisateur',
        operation_description='Cette vue vous permet de déconnecter un utilisateur en invalidant son token d\'accès.',
    )
    def post(self, request):
        authentication_classes = (JWTAuthentication,)
        permission_classes = (IsAuthenticated,)  # Permet uniquement aux utilisateurs authentifiés

        # Récupérer le token depuis l'en-tête Authorization
        auth_header = request.META.get('HTTP_AUTHORIZATION', "")
        if auth_header.startswith("Bearer "):
            user_token = auth_header.split(' ')[1]
        else:
            user_token = None

        if user_token:
            try:    
                # Vérifie si le token est un RefreshToken valide et le blacklist
                try:
                    refresh_token = RefreshToken(user_token)
                    refresh_token.blacklist()
                    response = Response({'message': 'Déconnexion réussie.'})
                    response.delete_cookie('access_token')
                    response.delete_cookie('refresh_token')
                    return response
                except TokenError:
                    return Response({'error': 'Le token fourni n\'est pas un RefreshToken valide.'}, status=status.HTTP_400_BAD_REQUEST)
                    
            except TokenError:
                # Si le token est invalide ou déjà blacklisté
                return Response({'error': 'Token invalide ou déjà blacklisté.'}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                # Gère les autres exceptions
                return Response({'error': 'Une erreur est survenue.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Si aucun token n'est fourni, l'utilisateur est déjà déconnecté
        return Response({'message': 'L\'utilisateur est déjà déconnecté.'}, status=status.HTTP_401_UNAUTHORIZED)


class get_pharmacie(APIView):
  authentication_classes = (JWTAuthentication,)
  permission_classes = (AllowAny,)

  def get(self,request):
        pharmacie=Pharmacie.objects.all()
        serializer=PharmacieSerializer(pharmacie,many=True)
        return Response(serializer.data)
        
class GetPharmacieGarde(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)

    def get(self, request):
        # Récupérer les pharmacies de garde
        pharmacies_de_garde = Pharmacie.objects.filter(degarde=True)
            # Serializer les pharmacies de garde si elles existent
        serializer = PharmacieSerializer(pharmacies_de_garde, many=True)
        return Response(serializer.data)
      
class get_Conseil(APIView):
    def get(self,request):
        conseils = Conseil.objects.all()
        serializer = ConseilSerializer(conseils, many=True)
        return Response(serializer.data)

class PasserCommandeClient(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (IsClientOrReadOnly,)

    def post(self, request):
        request_data = request.data
        clt=Client.objects.get(user=request.user.pk)
        request_data['client'] = clt
        
        if clt.num_pharmacie:
            pharmacie = Pharmacie.objects.filter(num_pharmacie= clt.num_pharmacie).first()
            if pharmacie is None:
                return Response({"detail": "Désolé, votre pharmacie est introuvable."}, status=status.HTTP_400_BAD_REQUEST)
            request_data['pharmacie_id'] = pharmacie
        else:
            return Response({"detail": "Désolé, veuillez d'abord renseigner le code de votre pharmacie."}, status=status.HTTP_400_BAD_REQUEST)
            
        commande = Commande.objects.create(**request_data)
        serializer =CommandeSerializer(commande, many=False)
        
        notification = Notification.objects.create(
            title = "Nouvelle commande",
            message = "Vous avez reçu une nouvelle commande.",
            type = "commande",
            data_id = commande.pk,
            metadata = CommandeSerializer(commande, many=False).data,
            is_read = False,
            user_type = 'pharmacie',
            user_id = pharmacie.pk
        )

        # SEND SMS
        if pharmacie.numero_contact_pharmacie:
            send_sms_jetfy(pharmacie.numero_contact_pharmacie, notification.message)
        
        if pharmacie.firebase_token:
            # Sérialisez l'instance de notification
            send_notification(notification, pharmacie.firebase_token)
        
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def get(self, request):
        
        commandes = Commande.objects.filter(client=request.user.client_user)
        serializer = CommandeSerializer(commandes, many=True)
        return Response(serializer.data)
    
class GestionCommandeDetailClient(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = [IsClientOrReadOnly]

    def get_object(self, pk):
        try:
            return Commande.objects.get(pk=pk)
        except Commande.DoesNotExist:
            raise Http404
    
    def put(self, request, pk):
        commande = self.get_object(pk)
        
        # Vérifier si la commande peut être modifiée
        if commande.en_attente:
            clt=Client.objects.get(user=request.user.pk)
            request.data['client'] = clt

            if request.data['pharmacie_id']:
                phmacie =Pharmacie.objects.get(pk=request.data['pharmacie_id'])
                if phmacie:
                    request.data['pharmacie_id'] = phmacie

            for cle, valeur in request.data.items():
                setattr(commande, cle, valeur)

            commande.save()

            serializer = CommandeSerializer(commande)
           

            return Response(serializer.data)
            # return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"detail": "La commande ne peut pas être modifiée car elle n'est plus en attente."}, status=status.HTTP_400_BAD_REQUEST)
        

    def delete(self, request, pk):
        commande = self.get_object(pk)
        
        # Vérifier si la commande peut être supprimée
        if commande.en_attente:
            commande.delete()
            return Response({"detail": "Commande supprimée avec succès"}, status=status.HTTP_200_OK)
        else:
            return Response({"detail": "La commande ne peut pas être supprimée car elle n'est plus en attente."}, status=status.HTTP_400_BAD_REQUEST)    

class CommandesPharmacietous(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = [IsPharmacieCanModifyCommande]
    
    def get(self, request):
        commandes = Commande.objects.filter(pharmacie_id=request.user.pharmacie_user.pk).order_by('-pk')
        
        
        serializer = CommandeSerializer(commandes, many=True)
        return Response(serializer.data)
        

class PharmacieDetail(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = [IsPharmacieCanModifyCommande]

    def get_object(self, pk):
        try:
            return Commande.objects.get(pk=pk)
        except Commande.DoesNotExist:
            raise Http404

    def get(self, request, pk):
        commande = self.get_object(pk)
        serializer = CommandeSerializer(commande)
        return Response(serializer.data)

    def put(self, request, pk):
        commande = self.get_object(pk)
        request.data['pharmacie_id'] =  commande.pharmacie_id.pk
        request.data['client'] =  commande.client.pk
    
        client = commande.client
        
        if request.data.get('statut')=='livree' and not commande.en_attente:
            return Response({"detail": "La commande a déjà été validée."}, status=status.HTTP_400_BAD_REQUEST)
        
        if request.data.get('facture') and request.data.get('items'):
            print("Welcome")
            # Valider la commande
            commande.en_attente = False
            commande.statut = 'traite'
            commande.save()

            items = request.data.get('items')
            if len(items) < 1:
                return Response({"detail": "La liste des médicaments est vide."}, status=status.HTTP_400_BAD_REQUEST)

            invoice= Invoice.objects.filter(commande = commande.pk).first()
            if not invoice:
                # Register Invoice
                invoice = Invoice.objects.create(
                    client = commande.client,
                    commande = commande,
                    due_date = request.data.get('date_echance'), 
                    facture = request.data.get('facture'),
                    status ='impaye',
                    reference = generate_reference(16)
                )
            else:
                if request.data.get('date_echance'):
                    invoice.due_date = request.data.get('date_echance')
                    invoice.save()
                if request.data.get('facture'):
                    invoice.facture = request.data.get('facture')
                    invoice.save()

                # Supprimer les ancie
                InvoiceItem.objects.filter(invoice=invoice.pk).delete()
            
            # register Items
            for item in items:
                InvoiceItem.objects.create(
                    invoice = invoice,
                    description = item["description"],  # Description du medicament
                    quantity = item["quantity"],
                    unit_price = item["unit_price"]
                )

            # Calcul du montant total
            invoice.calculate_total()


            notification = Notification.objects.create(
                title = "Nouvelle réponse pharmacie",
                message = "Vous avez reçu une réponse à votre commande #"+ str(commande.pk),
                type = "commande",
                data_id = commande.pk,
                metadata = CommandeSerializer(commande, many=False).data,
                is_read = False,
                user_type = 'client',
                user_id = client.pk
            )

            # SEND SMS
            if client.phone:
                send_sms_jetfy(client.phone, notification.message)
                        
            if client.firebase_token:
                send_notification(notification, client.firebase_token)
            
        if request.data.get('statut'):
            commande.statut = request.data.get('statut')
            if request.data.get('statut') == "termine":
                commande.termine = True
                
            if request.data.get('statut') == "annulee":
                commande.en_attente = False
                commande.termine = True
                
                notification = Notification.objects.create(
                    title = "Commande rejetée",
                    message = "Nous sommes dans le regret de vous annoncer que nous ne sommes pas à mesure de répondre favorablement à votre commande numero #"+str(commande.pk),
                    type = "commande",
                    data_id = commande.pk,
                    metadata = CommandeSerializer(commande, many=False).data,
                    is_read = False,
                    user_type = 'client',
                    user_id = client.pk
                )

                # SEND SMS
                if client.phone:
                    send_sms_jetfy(client.phone, notification.message)
                            
                if client.firebase_token:
                    send_notification(notification, client.firebase_token)
            
                
            commande.save()
                
        return Response(CommandeSerializer(commande, many=False).data)

    def delete(self, request, pk):
        customer = self.get_object(pk)
        customer.delete()
        return Response({"detail": "Utilisateur supprimée avec succès"}, status=status.HTTP_200_OK)        
    
class ConseilDetail(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = [IsPharmacieCanModifyCommande]
    
    def get_object(self, pk):
        try:
            return Conseil.objects.get(pk=pk)
        except Conseil.DoesNotExist:
            raise Http404

    
    def get(self, request, pk):
        conseil = self.get_object(pk)
        serializer = ConseilSerializer(conseil)
        return Response(serializer.data)
    
    def post(self, request):
        phar=Pharmacie.objects.get(user=request.user.pk)
        request.data['pharmacie'] = phar.pk
        serializer = ConseilSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    
    def put(self, request, pk):
        
        conseil = self.get_object(pk)
        phar=Pharmacie.objects.get(user=request.user.pk)
        request.data['pharmacie'] = phar.pk
        serializer = ConseilSerializer(conseil, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    
    def delete(self, request, pk):
        conseil = self.get_object(pk)
        conseil.delete()
        return Response({"detail": "Conseil supprimé avec succès"}, status=status.HTTP_200_OK)    
    
class RechercheList(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = [IsPharmacieOrClient]
    
    def get(self, request):
        if request.user.is_pharmacie == True:
            # recherches = Recherche.objects.filter(Q(pharmacie_id=request.user.pharmacie_user.pk) | Q(en_attente=True)).order_by("-pk")
            recherches = Recherche.objects.filter(pharmacie_id=request.user.pharmacie_user.pk).order_by("-pk")
        else:
            recherches = Recherche.objects.filter(client=request.user.client_user.pk).order_by("-pk")
    
        serializer = RechercheSerializer(recherches, many=True)
        return Response(serializer.data)

    def post(self, request):
        client = request.user.client_user

        if not client.num_pharmacie:
            return Response({"detail": "Désolé, vous n'êtes affilié a aucune pharmacie."}, status=status.HTTP_403_FORBIDDEN)
        
        pharmacie=Pharmacie.objects.filter(num_pharmacie=client.num_pharmacie).first()

        if  not pharmacie:
            return Response({"detail": "Désolé, aucune pharmacie trouvée avec votre identifiant pharmacie."}, status=status.HTTP_400_BAD_REQUEST)
            
        request.data["client"]= client
        request.data["pharmacie_id"]= pharmacie
                
        recherche = Recherche.objects.create(**request.data)
        
        serializer = RechercheSerializer(recherche, many=False)
        
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class RechercheDetail(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = [IsPharmacieOrClient]
    
    def get_object(self, pk):
        try:
            return Recherche.objects.get(pk=pk)
        except Recherche.DoesNotExist:
            raise Http404
            #return Response({"detail":"Recherche introuvable"}, status=status.HTTP_404_NOT_FOUND)

    def get(self, request, pk):
        recherche = self.get_object(pk)
        
        # Verifier si pharmacie
        if request.user.is_pharmacie == True :
            # Verifier si recherche traitée par d'autre pharmacie
            if recherche.en_attente == False and recherche.pharmacie_id.pk != request.user.pharmacie_user.pk:
                return Response({"detail":"Désolé, vous ne pouvez accéder à cette recherche"}, status=status.HTTP_403_FORBIDDEN)
        else:
            if recherche.client.pk != request.user.client_user.pk:
                return Response({"detail":"Désolé, vous ne pouvez accéder à cette recherche"}, status=status.HTTP_403_FORBIDDEN)
        serializer = RechercheSerializer(recherche)
        return Response(serializer.data)

    def put(self, request, pk):
        recherche = self.get_object(pk)
        
        request.data['client'] = recherche.client
        request.data['recherche'] = recherche
        if recherche.terminer == True:
            return Response({"detail":"Cette recherche est déjà terminée"}, status=status.HTTP_400_BAD_REQUEST)
        if request.user.is_pharmacie == True :
            if not('facture' in request.data) or not(request.data['facture']):
                return Response({"detail":"La facture est obligatoire"}, status=status.HTTP_400_BAD_REQUEST)
            
            if recherche.en_attente == False or recherche.statut == "traite":
                return Response({"detail":"Desolé, cette recherche a dejà été traitée"}, status=status.HTTP_400_BAD_REQUEST)
            
            
            orderData = { 'nom_medicament': recherche.nom_medicament}
            orderData['nom_medicament'] = recherche.nom_medicament
            orderData['ordornance'] = recherche.ordonnance
            orderData['quantite'] = recherche.quantite
            orderData['recherche'] = recherche
            orderData['client'] = recherche.client
            orderData['pharmacie_id'] = request.user.pharmacie_user

            commande = Commande.objects.create(**orderData)

            if request.data.get('facture') and request.data.get('items'):
                # Valider la commande
                commande.en_attente = False
                commande.statut = 'traite'
                commande.save()

                items = request.data.get('items')
                if len(items) < 1:
                    return Response({"detail": "La liste des médicaments est vide."}, status=status.HTTP_400_BAD_REQUEST)

                invoice= Invoice.objects.filter(commande = commande.pk).first()
                if not invoice:
                    # Register Invoice
                    invoice = Invoice.objects.create(
                        client = commande.client,
                        commande = commande,
                        due_date = request.data.get('date_echance'), 
                        facture = request.data.get('facture'),
                        status ='impaye',
                        reference = generate_reference(16)
                    )
                else:
                    if request.data.get('date_echance'):
                        invoice.due_date = request.data.get('date_echance')
                        invoice.save()
                    if request.data.get('facture'):
                        invoice.facture = request.data.get('facture')
                        invoice.save()

                    # Supprimer les ancie
                    InvoiceItem.objects.filter(invoice=invoice.pk).delete()
                
                # register Items
                for item in items:
                    InvoiceItem.objects.create(
                        invoice = invoice,
                        description = item["description"],  # Description du medicament
                        quantity = item["quantity"],
                        unit_price = item["unit_price"]
                    )

                # Calcul du montant total
                invoice.calculate_total()

            #Mise à jour de la recherche
            recherche.statut = "traite"
            recherche.facture = request.data.get('facture')
            recherche.en_attente = False
            recherche.pharmacie_id = request.user.pharmacie_user
            recherche.save()
            
            # Start Notification
            notification = Notification.objects.create(
                title = "Vous avez reçu une reponse à votre récherche",
                message = "Une pharmacie à repondu à votre recherche de médicament numéro #"+ str(recherche.pk),
                type = "recherche",
                data_id = recherche.pk,
                metadata = RechercheSerializer(recherche, many=False).data,
                is_read = False,
                user_type = 'client',
                user_id = recherche.client.pk
            )

            # Send Notification
            if recherche.client.phone:
                # Send SMS
                send_sms_jetfy(recherche.client.phone, notification.message)
                
            if recherche.client.firebase_token:
                send_notification(notification, recherche.client.firebase_token)

            # End Notification

            serializer = RechercheSerializer(recherche, many=False)
            return Response(serializer.data)
            
        if request.user.is_pharmacie != True :
            if recherche.client.pk != request.user.client_user.pk:
                return Response({"detail":"Désolé, vous ne pouvez effectuer à cette action"}, status=status.HTTP_403_FORBIDDEN)
            
            # Si le champ terminer est absent ou faux, renvoyer une erreur
            if  not('terminer' in request.data) or request.data['terminer'] == False:
                return Response({"detail":"Requete incorrecte"}, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Sinon, terminer la recherche
                recherche.en_attente = False
                recherche.statut = "termine"
                recherche.save()
                
                # Mettre à jour le statut de la commande
                commande = Commande.objects.get(pk=recherche.commande.pk)
                commande.statut = "termine"
                commande.save()
        
        for key in request.data:
            setattr(recherche, key, request.data[key])
        recherche.save()

        serializer = RechercheSerializer(recherche, many=False)
        return Response(serializer.data)

    def delete(self, request, pk):
        recherche = self.get_object(pk)
        if request.user.is_pharmacie != True :
            if recherche.client.pk != request.user.client_user.pk:
                return Response({"detail":"Désolé, vous ne pouvez effectuer à cette action"}, status=status.HTTP_403_FORBIDDEN)
        recherche.delete()
        return Response({"detail": "Recherche supprimée avec succès"},status=status.HTTP_200_OK)


class RechercheComamnde(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = [IsPharmacieOrClient]
    
    def get_object(self, pk):
        try:
            return Recherche.objects.get(pk=pk)
        except Recherche.DoesNotExist:
            raise Http404

    def get(self, request, pk):
        recherche = self.get_object(pk)

        if recherche.en_attente == True:
            return Response({"detail":"Cette recherche n'a pas encore été traitée"}, status=status.HTTP_403_FORBIDDEN)

        commande = Commande.objects.filter(recherche=recherche.pk).first()
        serializer = CommandeSerializer(commande)
        return Response(serializer.data)

class NotificationList(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)
    
    def get(self, request):
        if request.user:
            if request.user.is_pharmacie != True:
                notifications = Notification.objects.filter(user_type='client').filter(user_id=request.user.client_user.pk).order_by("-pk")
            else:
                notifications = Notification.objects.filter(user_type='pharmacie').filter(user_id=request.user.pharmacie_user.pk).order_by("-pk")
            
        serializer = NotificationSerializer(notifications, many=True)
        return Response(serializer.data)

    def post(self, request):
        data=request.data
        if request.user.is_pharmacie != True:
            data['user_type'] = 'client'
            data['user_id'] = request.user.client_user.pk
        else:
            data['user_type'] = 'pharmacie'
            data['user_id'] = request.user.pharmacie_user.pk
            
            
        serializer = NotificationSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class NotificationDetail(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)
    
    def get_object(self, pk):
        try:
            return Notification.objects.get(pk=pk)
        except Notification.DoesNotExist:
            raise Http404
        

    def get(self, request, pk):
        notification = self.get_object(pk)
        serializer = NotificationSerializer(notification)
        return Response(serializer.data)

    def put(self, request, pk):        
        notification = self.get_object(pk)
        
        # Mettre à jour les attributs de l'objet avec les valeurs fournies dans le dictionnaire
        for key in request.data:
            setattr(notification, key, request.data[key])
            
        # Enregistrer l'objet mis à jour dans la base de données
        notification.save()

        # notification.update(**data)
        return Response(NotificationSerializer(notification,many=False).data)

    def delete(self, request, pk):
        notification = self.get_object(pk)
        notification.delete()
        return Response({'detail': "Notification supprimée"},status=status.HTTP_200_OK)


class clientInvoices(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)

    def get(self, request):
        clt=Client.objects.get(user=request.user.pk)
        invoices = Invoice.objects.filter(client=clt.pk)
        serializer = InvoiceSerializer(invoices, many=True)
        return Response(serializer.data)

class orderInvoices(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)


    def get_order(self, pk):
        try:
            return Commande.objects.get(pk=pk)
        except Commande.DoesNotExist:
            raise Http404

    def get(self, request, commande):
        commande = self.get_order(commande)

        invoices = Invoice.objects.filter(commande=commande.pk)
        serializer = InvoiceSerializer(invoices, many=True)
        return Response(serializer.data)



class pharmacieInvoices(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)


    def get(self, request):
        phmcie =Pharmacie.objects.get(user=request.user.pk)

        invoices = Invoice.objects.filter(commande__pharmacie_id__id=phmcie.pk).order_by('-pk')
        serializer = InvoiceSerializer(invoices, many=True)
        return Response(serializer.data)
    
class initiate_payment(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)

    def post(self, request):
        clt=Client.objects.get(user=request.user.pk)
        print("clt : ", clt)
        print("request.data : ", request.data.get("invoice"))
        invoice = Invoice.objects.filter(pk=request.data.get("invoice")).first()
        print("invoice : ", invoice)
        if not invoice:
            return Response("Facture introuvable", status=status.HTTP_404_NOT_FOUND)
        
        if invoice.status == "payee":
            return Response("Facture deja payée", status=status.HTTP_403_FORBIDDEN)
        

        if invoice.client.pk != clt.pk :
            return Response({"detail":"Vous ne pouvez pas payer cette facture"}, status=status.HTTP_400_BAD_REQUEST)
        
        print("invoice : ", invoice)
        payment = InvoicePayment.objects.filter(invoice=invoice.pk).first()
        print("payment : ", payment)
        if not payment:
            payment = InvoicePayment.objects.create(
                reference = generate_reference(32),
                invoice = invoice,
                type_transaction = "collecte",
                amount_total = invoice.total,
                status = "initie",
                currency = "XOF"
            )
            payment.save()

        serializer = InvoicePaymentSerializer(payment)
        return Response(serializer.data)


class InvoiceListCreateAPIView(generics.ListCreateAPIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)

    queryset = Invoice.objects.prefetch_related('items')
    serializer_class = InvoiceSerializer

class InvoiceRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)

    queryset = Invoice.objects.prefetch_related('items')
    serializer_class = InvoiceSerializer

class InvoiceregisterRecuCode(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)

    def get_invoice(self, pk):
        try:
            return Invoice.objects.get(pk=pk)
        except Invoice.DoesNotExist:
            raise Http404

    def put(self, request, pk):
        invoice =self.get_invoice(pk)

        if not request.data["code_recu"]:
            return Response({"detail":"Désolé, le code du reçu pharmacie est réquis"}, status=status.HTTP_400_BAD_REQUEST)

        invoice.code_recu = request.data["code_recu"]
        invoice.save()

        serializer = InvoiceSerializer(invoice)
        return Response(serializer.data)

class  sendNotification(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)

    def post(self, request):
        # Start Notification
        notification = Notification.objects.create(
            title = "Vous avez reçu nouvelle notification",
            message = "Vous avez recu une nouvelle notification test", 
            type = "test",
            data_id = 0,
            metadata = "test data",
            is_read = False,
            user_type = 'client',
            user_id = request.user.client_user.pk
        )

        # SEND SMS
        if request.user.client_user.phone:
            data = send_sms_jetfy(request.user.client_user.phone, notification.message)
            print("SMS data : ", data)
        
        if request.user.client_user.firebase_token:
            send_notification(notification, request.user.client_user.firebase_token)

        return Response({"detail":"Notification envoyée"}, status=status.HTTP_200_OK)

###################  Phamacy wallet #########################

# *****************  Admin  *********************************
def get_or_create_wallet_pharmacie(pharmacie_id):
        try:
            pharmacie = Pharmacie.objects.get(pk=pharmacie_id)
        except Pharmacie.DoesNotExist:
            raise ValueError("Pharmacie non trouvée")
            
        wallet, created = WalletPharmacie.objects.get_or_create(
            pharmacie=pharmacie,
            defaults={'balance': 0.00, 'old_balance': 0.00}
        )
        
        if created:
            WalletPharmacieHistory.objects.create(
                wallet=wallet,
                action_type='depot',
                label='Création du wallet',
                amount=0.00,
                new_balance=0.00
            )
            
        return wallet, created

#  Crediter  ou  debiter  le  solde  de  la  pharmacie
class WalletTransactionView(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)

    @transaction.atomic
    def post(self, request, pharmacie_id):
        try:           
            wallet, is_new = get_or_create_wallet_pharmacie(pharmacie_id)
            
            # Validation des données
            serializer = TransactionSerializer(data=request.data)
            
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            data = serializer.validated_data
            amount = Decimal(str(data['amount']))
            
            if data['action_type'] == 'retrait' and amount > wallet.balance:
                return Response(
                    {"error": "Solde insuffisant"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Mise à jour du solde
            wallet.old_balance = wallet.balance
            if data['action_type'] == 'depot':
                wallet.balance += amount
            else:
                wallet.balance -= amount
            wallet.save()
            
            # Création de l'historique
            WalletPharmacieHistory.objects.create(
                wallet=wallet,
                action_type=data['action_type'],
                label=data.get('label'),
                amount=amount,
                user=request.user,
                new_balance=wallet.balance
            )
            
            return Response({
                "message": f"{data['action_type']} effectué avec succès",
                "new_balance": wallet.balance
            }, status=status.HTTP_200_OK)
            
        except WalletPharmacie.DoesNotExist:
            return Response(
                {"error": "Wallet non trouvé"}, 
                status=status.HTTP_404_NOT_FOUND
            )

class WalletPharmacieView(APIView):
    def get(self, request, pharmacie_id):
        wallet, is_new = get_or_create_wallet_pharmacie(pharmacie_id)
        serializer = WalletPharmacieSerializer(wallet)
        return Response(serializer.data)


class  WalletHistoryView(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)
    
    def get(self, request, pharmacie_id):
        wallet = WalletPharmacie.objects.get(pharmacie=pharmacie_id)
        history = WalletPharmacieHistory.objects.filter(wallet=wallet).order_by('-created_at')
        serializer = WalletPharmacieHistorySerializer(history, many=True)
        return Response(serializer.data)


# *****************  Pharmacie  *********************************
class  PharmacyWallet(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)

    def get(self, request):
        if request.user.is_pharmacie != True:
            return Response("Autorisation refusée", status=status.HTTP_403_FORBIDDEN)

        pharmacie_id = request.user.pharmacie_user.pk
        wallet, is_new = get_or_create_wallet_pharmacie(pharmacie_id)

        serializer = WalletPharmacieSerializer(wallet)
        return Response(serializer.data)


class  PharmacyWalletHistory(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)
    
    def get(self, request):
        if request.user.is_pharmacie != True:
            return Response("Autorisation refusée", status=status.HTTP_403_FORBIDDEN)

        pharmacie_id = request.user.pharmacie_user.pk
        wallet, is_new = get_or_create_wallet_pharmacie(pharmacie_id)
        history = WalletPharmacieHistory.objects.filter(wallet=wallet).order_by('-created_at')
        serializer = WalletPharmacieHistorySerializer(history, many=True)
        return Response(serializer.data)



class BroadcastMessageAPIView(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)
    
    parser_classes = [MultiPartParser]  # Pour accepter les fichiers multipart/form-data

    
    def post(self, request, *args, **kwargs):
        # Récupérer les données envoyées dans la requête
        fichier_excel = request.FILES.get('fichier')  # Le fichier Excel
        message_generique = request.data.get('message')  # Le message à diffuser

        if not fichier_excel or not message_generique:
            return Response({"error": "Veuillez fournir un fichier Excel et un message."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Lire les noms et numéros à partir du fichier Excel
            wb = openpyxl.load_workbook(fichier_excel)
            sheet = wb.active

            # Récupérer les données : première colonne = nom, deuxième colonne = numéro
            clients = [(row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2, max_col=2)]

            # Envoyer le message personnalisé à chaque client
            resultats = []
            for nom, numero in clients:
                if numero and nom:  # Vérifier que le numéro et le nom ne sont pas vides
                    try:
                        # Construire le message personnalisé
                        prefixe = "Mme/Mlle/Mr"  # Vous pouvez personnaliser cela
                        message_personnalise = f"{prefixe} {nom}, {message_generique}"

                        # Envoyer le message via Twilio
                        data = send_sms_jetfy(numero, message_personnalise)
                        resultats.append(data)
                    except Exception as e:
                        resultats.append({"nom": nom, "numero": numero, "statut": f"échec: {str(e)}"})

            return Response({"resultats": resultats}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class BroadcastMessageWithConditionAPIView(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)
    
    parser_classes = [MultiPartParser]  # Pour accepter les fichiers multipart/form-data

    
    def post(self, request, *args, **kwargs):
        # Récupérer les données envoyées dans la requête
        fichier_excel = request.FILES.get('fichier')  # Le fichier Excel
        message_generique = request.data.get('message')  # Le message à diffuser

        if not fichier_excel or not message_generique:
            return Response({"error": "Veuillez fournir un fichier Excel et un message."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Lire les noms et numéros à partir du fichier Excel
            wb = openpyxl.load_workbook(fichier_excel)
            sheet = wb.active

            clients = []

            # Récupérer les données : première colonne = nom, deuxième colonne = numéro
            #clients = [(row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2, max_col=2)]

            for row in sheet.iter_rows(min_row=2, max_col=4):

                # Cas 1 : Colonne Beneficiaire est vide
                if row[0].value and not(row[1].value):
                    clients.append((row[0].value, row[3].value))
                

                # Cas 2 :  Colonnes Nom et beneficiaire sont pas vides
                if row[0].value and row[1].value:
                    if row[2].value and row[2].value < 18:
                        clients.append((row[0].value, row[3].value))
                    else:
                        clients.append((row[1].value, row[3].value))
                
                # Cas 3 : Colonnes Nom et beneficiaire sont vides
                if not(row[0].value) and not(row[1].value):
                    clients.append((row[0].value, row[3].value))

            print(clients)

            # Envoyer le message personnalisé à chaque client
            resultats = []
            for nom, numero in clients:
                if numero:  # Vérifier que le numéro et le nom ne sont pas vides
                    try:
                        # Construire le message personnalisé
                        prefixe = "Mme/Mlle/Mr"  # Vous pouvez personnaliser cela
                        if nom:
                            message_personnalise = f"{prefixe} {nom}, {message_generique}"
                        else:
                            message_personnalise = f"{message_generique}"

                        # Envoyer le message via Twilio
                        data = send_sms_jetfy(numero, message_personnalise)
                        resultats.append(data)
                    except Exception as e:
                        resultats.append({"nom": nom, "numero": numero, "statut": f"échec: {str(e)}"})

            return Response({"resultats": resultats}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ImageUploadView(APIView):
    authentication_classes = (JWTAuthentication,)
    permission_classes = (AllowAny,)
    
    parser_classes = (MultiPartParser, FormParser)
    
    def post(self, request):
        # Vérifier si on a des fichiers
        files = request.FILES.getlist('images')
        if not files:
            return Response(
                {'error': 'Aucune image fournie'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        uploaded_images = []
        errors = []
        
        # Utiliser une transaction pour garantir l'intégrité des données
        with transaction.atomic():
            for file in files:
                try:
                    # Upload sur S3
                    s3_url = upload_to_s3(file, file.name)
                    
                    uploaded_images.append({
                        'file': file.name,
                        'url': s3_url
                    })
                except Exception as e:
                    errors.append({
                        'file': file.name,
                        'error': str(e)
                    })
        
        # Préparer la réponse
        response_data = {
            'success': uploaded_images,
            'errors': errors
        }
        
        # Si certains uploads ont échoué mais pas tous
        if errors and uploaded_images:
            status_code = status.HTTP_207_MULTI_STATUS
        # Si tout a réussi
        elif uploaded_images and not errors:
            status_code = status.HTTP_201_CREATED
        # Si tout a échoué
        else:
            status_code = status.HTTP_500_INTERNAL_SERVER_ERROR
            
        return Response(response_data, status=status_code)
